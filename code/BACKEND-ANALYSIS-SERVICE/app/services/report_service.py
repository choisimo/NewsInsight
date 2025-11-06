from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
import json
import uuid
import io
import base64
from collections import defaultdict
from fastapi import BackgroundTasks
from fastapi.responses import FileResponse, StreamingResponse
from app.db import Report, SentimentAnalysis, TrendAnalysis
from app.schemas import ReportRequest, ReportResponse
import logging

# PDF/Excel generation (optional dependencies)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.charts.piecharts import Pie
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    
try:
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ReportService:
    """
    감성/트렌드 데이터 기반의 리포트를 생성/조회/다운로드하는 서비스.

    JSON 응답을 기본으로 하며, 선택적으로 PDF/Excel 형식으로 내보내기를 지원합니다.
    """

    def __init__(self, db: Session):
        """
        서비스 인스턴스 초기화.

        Args:
            db: SQLAlchemy 세션
        """
        self.db = db
    
    async def generate_report(self, request: ReportRequest, background_tasks: BackgroundTasks) -> ReportResponse:
        """
        요청된 타입/파라미터에 맞는 리포트를 생성하고 저장합니다.

        Args:
            request: 리포트 생성 요청(타입/제목/파라미터 포함)
            background_tasks: 백그라운드 태스크 큐(대용량 생성 시 활용 가능)

        Returns:
            생성된 리포트 응답 스키마
        """
        report_content = await self._generate_report_content(request)
        
        report = Report(
            title=request.title,
            report_type=request.report_type,
            content=json.dumps(report_content),
            parameters=json.dumps(request.parameters),
            created_by="system"
        )
        
        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)
        
        return ReportResponse(
            report_id=report.id,
            title=report.title,
            report_type=report.report_type,
            content=report_content,
            created_at=report.created_at,
            download_url=f"/api/v1/reports/{report.id}/download"
        )
    
    async def list_reports(self, report_type: Optional[str] = None, 
                          limit: int = 10, offset: int = 0) -> List[ReportResponse]:
        """
        리포트 목록을 유형/페이지네이션 기준으로 조회합니다.

        Args:
            report_type: 리포트 타입 필터
            limit: 최대 개수
            offset: 시작 오프셋

        Returns:
            ReportResponse 리스트
        """
        query = self.db.query(Report).filter(Report.is_active == True)
        
        if report_type:
            query = query.filter(Report.report_type == report_type)
        
        reports = query.offset(offset).limit(limit).all()
        
        return [
            ReportResponse(
                report_id=report.id,
                title=report.title,
                report_type=report.report_type,
                content=json.loads(report.content),
                created_at=report.created_at,
                download_url=f"/api/v1/reports/{report.id}/download"
            )
            for report in reports
        ]
    
    async def get_report(self, report_id: int) -> Optional[ReportResponse]:
        """
        단일 리포트 상세를 조회합니다.

        Args:
            report_id: 리포트 ID

        Returns:
            ReportResponse 또는 None
        """
        report = self.db.query(Report).filter(
            Report.id == report_id,
            Report.is_active == True
        ).first()
        
        if not report:
            return None
        
        return ReportResponse(
            report_id=report.id,
            title=report.title,
            report_type=report.report_type,
            content=json.loads(report.content),
            created_at=report.created_at,
            download_url=f"/api/v1/reports/{report.id}/download"
        )
    
    async def delete_report(self, report_id: int) -> bool:
        """
        리포트를 비활성(소프트 삭제) 처리합니다.

        Args:
            report_id: 리포트 ID

        Returns:
            성공 여부
        """
        report = self.db.query(Report).filter(Report.id == report_id).first()
        
        if not report:
            return False
        
        report.is_active = False
        self.db.commit()
        return True
    
    async def download_report(self, report_id: int, format: str = "json"):
        """
        리포트 다운로드 (JSON/PDF/Excel)
        """
        report = await self.get_report(report_id)
        
        if not report:
            return None
        
        if format == "json":
            return report.content
        elif format == "pdf" and PDF_AVAILABLE:
            return await self._generate_pdf(report)
        elif format == "excel" and EXCEL_AVAILABLE:
            return await self._generate_excel(report)
        else:
            return {"error": f"Format '{format}' not supported or libraries not installed"}
    
    def _generate_sentiment_insights(self, percentages: Dict[str, float], 
                                     change: float, total: int) -> List[str]:
        """
        감성 분석 인사이트 생성
        """
        insights = []
        
        # 긍정 비율 분석
        if percentages["positive"] > 50:
            insights.append(f"긍정적 의견이 {percentages['positive']:.1f}%로 절반 이상을 차지합니다.")
        elif percentages["negative"] > 50:
            insights.append(f"부정적 의견이 {percentages['negative']:.1f}%로 우세합니다.")
        
        # 변화 분석
        if abs(change) > 10:
            direction = "증가" if change > 0 else "감소"
            insights.append(f"긍정 감성이 기간 대비 {abs(change):.1f}% {direction}했습니다.")
        
        # 데이터 규모
        if total > 1000:
            insights.append(f"충분한 샘플 수({total}건)로 신뢰도가 높습니다.")
        elif total < 100:
            insights.append(f"샘플 수({total}건)가 적어 해석에 주의가 필요합니다.")
        
        return insights
    
    async def _generate_pdf(self, report: ReportResponse):
        """
        PDF 리포트 생성
        """
        if not PDF_AVAILABLE:
            return {"error": "PDF generation not available"}
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 제목
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30
        )
        story.append(Paragraph(report.title, title_style))
        story.append(Spacer(1, 12))
        
        # 생성 시간
        story.append(Paragraph(f"생성일시: {report.created_at.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # 요약
        if "summary" in report.content:
            story.append(Paragraph("요약", styles['Heading2']))
            story.append(Paragraph(report.content["summary"], styles['Normal']))
            story.append(Spacer(1, 12))
        
        # 주요 발견사항
        if "key_findings" in report.content:
            story.append(Paragraph("주요 발견사항", styles['Heading2']))
            for finding in report.content["key_findings"]:
                story.append(Paragraph(f"• {finding}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        # 통계 테이블
        if "sentiment_distribution" in report.content:
            story.append(Paragraph("감성 분포", styles['Heading2']))
            dist = report.content["sentiment_distribution"]["percentages"]
            data = [["감성", "비율"]]
            data.extend([[k.capitalize(), f"{v:.1f}%"] for k, v in dist.items()])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=report_{report.report_id}.pdf"}
        )
    
    async def _generate_excel(self, report: ReportResponse):
        """
        Excel 리포트 생성
        """
        if not EXCEL_AVAILABLE:
            return {"error": "Excel generation not available"}
        
        wb = Workbook()
        ws = wb.active
        ws.title = "리포트 요약"
        
        # 제목
        ws['A1'] = report.title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"생성일시: {report.created_at.strftime('%Y-%m-%d %H:%M')}"
        
        row = 4
        
        # 요약
        if "summary" in report.content:
            ws[f'A{row}'] = "요약"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = report.content["summary"]
            row += 2
        
        # 감성 분포
        if "sentiment_distribution" in report.content:
            ws[f'A{row}'] = "감성 분포"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            dist = report.content["sentiment_distribution"]["percentages"]
            ws[f'A{row}'] = "감성"
            ws[f'B{row}'] = "비율 (%)"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for sentiment, percentage in dist.items():
                ws[f'A{row}'] = sentiment.capitalize()
                ws[f'B{row}'] = round(percentage, 1)
                row += 1
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=report_{report.report_id}.xlsx"}
        )
    
    async def _generate_report_content(self, request: ReportRequest) -> Dict[str, Any]:
        """
        요청 타입에 따라 실제 리포트 콘텐츠를 생성합니다.

        Args:
            request: 리포트 생성 요청

        Returns:
            리포트 콘텐츠 사전(JSON 직렬화 가능)
        """
        if request.report_type == "sentiment":
            return await self._generate_sentiment_report(request)
        elif request.report_type == "trend":
            return await self._generate_trend_report(request)
        elif request.report_type == "summary":
            return await self._generate_summary_report(request)
        else:
            return {"error": "Unknown report type"}
    
    async def _generate_sentiment_report(self, request: ReportRequest) -> Dict[str, Any]:
        """
        실제 데이터 기반 감성 분석 리포트 생성
        """
        params = request.parameters
        start_date = params.get("start_date")
        end_date = params.get("end_date")
        
        # 날짜 기본값 설정
        if not end_date:
            end_date = datetime.now()
        elif isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            
        if not start_date:
            start_date = end_date - timedelta(days=30)
        elif isinstance(start_date, str):
            start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        
        # 데이터베이스에서 실제 감성 분석 결과 조회
        analyses = self.db.query(SentimentAnalysis).filter(
            SentimentAnalysis.created_at >= start_date,
            SentimentAnalysis.created_at <= end_date
        ).all()
        
        if not analyses:
            return {
                "type": "sentiment_analysis",
                "title": request.title,
                "generated_at": datetime.now().isoformat(),
                "period": {"start": start_date.isoformat(), "end": end_date.isoformat()},
                "summary": "데이터가 없습니다.",
                "data": {"total_analyzed": 0}
            }
        
        # 감성별 분류
        sentiment_counts = {"positive": 0, "negative": 0, "neutral": 0}
        sentiment_scores = []
        daily_data = defaultdict(lambda: {"positive": 0, "negative": 0, "neutral": 0, "total": 0})
        
        for analysis in analyses:
            sentiment_counts[analysis.sentiment_label] += 1
            sentiment_scores.append(analysis.sentiment_score)
            
            # 일별 집계
            date_key = analysis.created_at.strftime("%Y-%m-%d")
            daily_data[date_key][analysis.sentiment_label] += 1
            daily_data[date_key]["total"] += 1
        
        total = len(analyses)
        avg_score = sum(sentiment_scores) / total if total > 0 else 0
        
        # 백분율 계산
        percentages = {
            label: round((count / total) * 100, 2) if total > 0 else 0
            for label, count in sentiment_counts.items()
        }
        
        # 일별 추이 데이터
        daily_trend = [
            {
                "date": date,
                "positive": data["positive"],
                "negative": data["negative"],
                "neutral": data["neutral"],
                "total": data["total"]
            }
            for date, data in sorted(daily_data.items())
        ]
        
        # 감성 변화 분석
        if len(daily_trend) >= 2:
            first_day_ratio = (daily_trend[0]["positive"] / daily_trend[0]["total"]) if daily_trend[0]["total"] > 0 else 0
            last_day_ratio = (daily_trend[-1]["positive"] / daily_trend[-1]["total"]) if daily_trend[-1]["total"] > 0 else 0
            sentiment_change = round((last_day_ratio - first_day_ratio) * 100, 2)
        else:
            sentiment_change = 0
        
        return {
            "type": "sentiment_analysis",
            "title": request.title,
            "generated_at": datetime.now().isoformat(),
            "period": {"start": start_date.isoformat(), "end": end_date.isoformat()},
            "summary": f"기간 내 {total}건 분석, 평균 감성 점수: {avg_score:.3f}",
            "sentiment_distribution": {
                "counts": sentiment_counts,
                "percentages": percentages
            },
            "statistics": {
                "total_analyzed": total,
                "average_sentiment_score": round(avg_score, 3),
                "sentiment_change_percent": sentiment_change
            },
            "daily_trend": daily_trend,
            "key_insights": self._generate_sentiment_insights(
                percentages, sentiment_change, total
            )
        }
    
    async def _generate_trend_report(self, request: ReportRequest) -> Dict[str, Any]:
        """
        실제 데이터 기반 트렌드 분석 리포트 생성
        """
        params = request.parameters
        period = params.get("period", "weekly")
        entity = params.get("entity")
        
        # 최근 30일 트렌드 데이터 조회
        trends = self.db.query(TrendAnalysis).filter(
            TrendAnalysis.period == period,
            TrendAnalysis.analysis_date >= datetime.now() - timedelta(days=30)
        )
        
        if entity:
            trends = trends.filter(TrendAnalysis.entity == entity)
        
        trends = trends.order_by(TrendAnalysis.analysis_date.desc()).all()
        
        if not trends:
            return {
                "type": "trend_analysis",
                "title": request.title,
                "generated_at": datetime.now().isoformat(),
                "summary": "트렌드 데이터가 없습니다.",
                "data": {}
            }
        
        # 엔티티별 트렌드 분류
        entity_trends = defaultdict(list)
        for trend in trends:
            entity_trends[trend.entity].append(trend)
        
        trending_up = []
        trending_down = []
        stable = []
        
        for entity_name, entity_trend_list in entity_trends.items():
            if len(entity_trend_list) < 2:
                continue
                
            # 최신과 최초 비교
            recent = entity_trend_list[0]
            older = entity_trend_list[-1]
            
            sentiment_change = recent.sentiment_trend - older.sentiment_trend
            volume_change = recent.volume_trend - older.volume_trend
            
            trend_info = {
                "entity": entity_name,
                "sentiment_change": round(sentiment_change, 3),
                "volume_change": volume_change,
                "current_sentiment": round(recent.sentiment_trend, 3),
                "keywords": json.loads(recent.keywords) if recent.keywords else []
            }
            
            if sentiment_change > 0.1 or volume_change > 10:
                trending_up.append(trend_info)
            elif sentiment_change < -0.1 or volume_change < -10:
                trending_down.append(trend_info)
            else:
                stable.append(trend_info)
        
        # 정렬
        trending_up.sort(key=lambda x: x["sentiment_change"], reverse=True)
        trending_down.sort(key=lambda x: x["sentiment_change"])
        
        return {
            "type": "trend_analysis",
            "title": request.title,
            "generated_at": datetime.now().isoformat(),
            "period": period,
            "summary": f"총 {len(entity_trends)}개 엔티티 분석, {len(trending_up)}개 상승 트렌드",
            "trending_up": trending_up[:10],
            "trending_down": trending_down[:10],
            "stable": stable[:10],
            "total_entities_analyzed": len(entity_trends)
        }
    
    async def _generate_summary_report(self, request: ReportRequest) -> Dict[str, Any]:
        """
        종합 요약 리포트 생성 (감성 + 트렌드 통합)
        """
        # 감성 분석 데이터
        sentiment_report = await self._generate_sentiment_report(request)
        # 트렌드 분석 데이터  
        trend_report = await self._generate_trend_report(request)
        
        # 주요 발견사항 생성
        key_findings = []
        
        # 감성 분석 인사이트
        if "sentiment_distribution" in sentiment_report:
            percentages = sentiment_report["sentiment_distribution"]["percentages"]
            dominant_sentiment = max(percentages, key=percentages.get)
            key_findings.append(
                f"{dominant_sentiment.capitalize()} 감성이 {percentages[dominant_sentiment]}%로 가장 높습니다."
            )
            
        if "statistics" in sentiment_report:
            stats = sentiment_report["statistics"]
            if abs(stats.get("sentiment_change_percent", 0)) > 5:
                direction = "증가" if stats["sentiment_change_percent"] > 0 else "감소"
                key_findings.append(
                    f"긍정 감성이 {abs(stats['sentiment_change_percent'])}% {direction}했습니다."
                )
        
        # 트렌드 인사이트
        if "trending_up" in trend_report and trend_report["trending_up"]:
            top_entity = trend_report["trending_up"][0]
            key_findings.append(
                f"{top_entity['entity']}가 가장 큰 상승 트렌드를 보이고 있습니다."
            )
        
        if "trending_down" in trend_report and trend_report["trending_down"]:
            key_findings.append(
                f"{len(trend_report['trending_down'])}개 엔티티가 하락 트렌드를 보입니다."
            )
        
        return {
            "type": "summary",
            "title": request.title,
            "generated_at": datetime.now().isoformat(),
            "summary": "국민연금 감성 및 트렌드 종합 분석 리포트",
            "key_findings": key_findings,
            "sentiment_summary": {
                "total_analyzed": sentiment_report.get("statistics", {}).get("total_analyzed", 0),
                "average_score": sentiment_report.get("statistics", {}).get("average_sentiment_score", 0),
                "distribution": sentiment_report.get("sentiment_distribution", {})
            },
            "trend_summary": {
                "trending_up_count": len(trend_report.get("trending_up", [])),
                "trending_down_count": len(trend_report.get("trending_down", [])),
                "stable_count": len(trend_report.get("stable", [])),
                "top_trending": trend_report.get("trending_up", [])[:3]
            },
            "detailed_reports": {
                "sentiment": sentiment_report,
                "trend": trend_report
            }
        }